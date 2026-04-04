"""Export helpers for P190 NavConverter desktop surfaces."""

from __future__ import annotations

import csv
from pathlib import Path

from desktop.services.output_package_service import (
    OutputArtifact,
    discover_output_package_entries,
)


def _normalize_path(output_path: str, suffix: str = ".xlsx") -> Path:
    path = Path(output_path)
    if path.suffix.lower() != suffix:
        path = path.with_suffix(suffix)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def build_manifest_rows(path_a: str, path_b: str) -> list[dict]:
    rows: list[dict] = []
    for side, source_path in (("A", path_a), ("B", path_b)):
        for artifact in discover_output_package_entries(source_path):
            size_kb = artifact.path.stat().st_size / 1024 if artifact.exists else None
            rows.append({
                "side": side,
                "key": artifact.key,
                "label": artifact.label,
                "path": str(artifact.path),
                "exists": artifact.exists,
                "kind": artifact.kind,
                "note": artifact.note,
                "size_kb": size_kb,
            })
    return rows


def export_package_manifest(
    path_a: str,
    path_b: str,
    output_path: str,
) -> str:
    """Export a side-by-side package inventory as XLSX, with CSV fallback."""
    path = _normalize_path(output_path, ".xlsx")
    rows = build_manifest_rows(path_a, path_b)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return export_package_manifest_csv(path_a, path_b, str(path.with_suffix(".csv")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"

    title_font = Font(name="Pretendard", size=14, bold=True, color="1B365D")
    header_font = Font(name="Pretendard", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="2A2A4A"))

    summary = {
        "A": sum(1 for row in rows if row["side"] == "A" and row["exists"]),
        "B": sum(1 for row in rows if row["side"] == "B" and row["exists"]),
        "A_total": sum(1 for row in rows if row["side"] == "A"),
        "B_total": sum(1 for row in rows if row["side"] == "B"),
    }

    ws.merge_cells("A1:H1")
    ws["A1"] = "P190 Package Manifest"
    ws["A1"].font = title_font
    ws["A2"] = "Style A ready"
    ws["B2"] = summary["A"]
    ws["C2"] = "/"
    ws["D2"] = summary["A_total"]
    ws["F2"] = "Style B ready"
    ws["G2"] = summary["B"]
    ws["H2"] = "/"
    ws["I2"] = summary["B_total"]
    ws.append([])
    ws.append([
        "Side", "Key", "Label", "Exists", "Kind", "Path", "Note", "Size KB",
    ])

    header_row = 4
    for col in range(1, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, row in enumerate(rows, header_row + 1):
        values = [
            row["side"],
            row["key"],
            row["label"],
            "Yes" if row["exists"] else "No",
            row["kind"],
            row["path"],
            row["note"],
            "" if row["size_kb"] is None else round(row["size_kb"], 1),
        ]
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = [8, 16, 24, 10, 12, 46, 34, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A5"
    wb.save(path)
    return str(path)


def export_package_manifest_csv(
    path_a: str,
    path_b: str,
    output_path: str,
) -> str:
    """CSV fallback for the package manifest."""
    path = _normalize_path(output_path, ".csv")
    rows = build_manifest_rows(path_a, path_b)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["P190 Package Manifest"])
        writer.writerow(["Style A ready", sum(1 for row in rows if row["side"] == "A" and row["exists"])])
        writer.writerow(["Style B ready", sum(1 for row in rows if row["side"] == "B" and row["exists"])])
        writer.writerow([])
        writer.writerow(["Side", "Key", "Label", "Exists", "Kind", "Path", "Note", "Size KB"])
        for row in rows:
            writer.writerow([
                row["side"],
                row["key"],
                row["label"],
                "Yes" if row["exists"] else "No",
                row["kind"],
                row["path"],
                row["note"],
                "" if row["size_kb"] is None else round(row["size_kb"], 1),
            ])
    return str(path)

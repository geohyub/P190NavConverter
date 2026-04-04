"""Tests for P190 package manifest exports."""

from __future__ import annotations

from pathlib import Path

from desktop.services.export_service import (
    export_package_manifest,
    export_package_manifest_csv,
)


def _make_package_file(tmp_path: Path, stem: str) -> Path:
    output = tmp_path / f"{stem}.p190"
    output.write_text("P190 SAMPLE\n", encoding="utf-8")
    (tmp_path / f"{stem}_QC_Report.txt").write_text("QC OK\n", encoding="utf-8")
    (tmp_path / f"{stem}_Track_Plot.png").write_bytes(b"PNG")
    (tmp_path / f"{stem}_RadEx_Geometry.tsv").write_text("ffid\tx\ty\n", encoding="utf-8")
    (tmp_path / f"{stem}_RadEx_Geometry_Aligned.txt").write_text("aligned\n", encoding="utf-8")
    (tmp_path / f"{stem}_FFID_Map.tsv").write_text("orig\tp190\n", encoding="utf-8")
    (tmp_path / f"{stem}_Feathering_Report.txt").write_text("feathering\n", encoding="utf-8")
    return output


def test_export_package_manifest_creates_readable_file(tmp_path: Path):
    path_a = _make_package_file(tmp_path, "style_a")
    path_b = _make_package_file(tmp_path, "style_b")
    output = tmp_path / "manifest.xlsx"

    saved = Path(export_package_manifest(str(path_a), str(path_b), str(output)))

    assert saved.exists()
    assert saved.suffix in {".xlsx", ".csv"}
    content = saved.read_text(encoding="utf-8-sig") if saved.suffix == ".csv" else ""
    if saved.suffix == ".csv":
        assert "P190 Package Manifest" in content
        assert "Style A ready" in content
        assert "Side,Key,Label" in content
    else:
        import openpyxl

        wb = openpyxl.load_workbook(saved)
        ws = wb.active
        assert ws["A1"].value == "P190 Package Manifest"
        assert ws["A2"].value == "Style A ready"
        assert ws["F2"].value == "Style B ready"
        assert ws.max_row >= 5


def test_export_package_manifest_csv_contains_expected_rows(tmp_path: Path):
    path_a = _make_package_file(tmp_path, "style_a")
    path_b = _make_package_file(tmp_path, "style_b")
    output = tmp_path / "manifest.csv"

    saved = Path(export_package_manifest_csv(str(path_a), str(path_b), str(output)))
    text = saved.read_text(encoding="utf-8-sig")

    assert saved.exists()
    assert saved.suffix == ".csv"
    assert "P190 Package Manifest" in text
    assert "Style A ready" in text
    assert "Style B ready" in text
    assert "Side,Key,Label,Exists,Kind,Path,Note,Size KB" in text
